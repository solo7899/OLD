from openpyxl import load_workbook
import os
from datetime import timedelta, date, datetime

a = os.path.dirname(__file__)
os.chdir(a)
days = timedelta(days=1)

wb = load_workbook('test/Test1.xlsx')
# print(wb.sheetnames)
a = 1
starting = datetime(year=2010, month=1, day=1)
data_list = []
ws = wb['Sheet1']
stop = ''
row = 1
while True:
    cell = ws['B'+ str(row)].value
    if not cell:
        break
    stop =  cell
    row += 1

while starting <= stop:
    data_list.append([0, starting, 0])
    starting += days
    
for index, i in enumerate(data_list):
    if ws['A'+ str(index+1)]:
        break
    cell_date = ws['B' + str(index+1)].value
    if cell_date == i[1]:
        i[0] = ws['A'+ str(index+1)].value
        i[2] = ws['C' + str(index+1)].value

data_list.sort(key=lambda x: x[1])
    


# mylist.sort()
       
wb.create_sheet(title='Result')
ws = wb['Result']

for index, item in enumerate(data_list):
    ws['A' + str(index+1)].value = item[0]
    ws['B' + str(index+1)].value = item[1].strftime('%m/%d/%Y')
    ws['C' + str(index+1)].value = item[2]
    
wb.save('test/Test1.xlsx')